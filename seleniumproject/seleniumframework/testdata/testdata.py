import openpyxl
file = openpyxl.load_workbook("/Users/sagarkamthane/Downloads/exceldemo.xlsx")
sheet = file.active


class HomePageTestData:

    signuppagetestdata = [{"firstname" : "dinesh", "lastname" : "more", "email" : "dinesh.more1@gmail.com"}]
    logintestdata = [{"id": "dinesh.more@gmail.com", "pass": "Dinesh@1312"},]

    @staticmethod
    def get_test_data(testcasename):
        Dict = {}
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == testcasename:
                for col in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
        return [Dict]
