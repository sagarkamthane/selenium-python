import openpyxl
file = openpyxl.load_workbook("/Users/sagarkamthane/Downloads/exceldemo.xlsx")
sheet = file.active
cell = sheet.cell(row = 2, column= 1)
print(cell.value)

sheet.cell(row = 3, column= 1).value = "sagar"
print(sheet.cell(row = 3, column= 1).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A2'].value)
print("="*50)
Dict = {}
for row in range(1, sheet.max_row+1 ):
    if sheet.cell(row= row, column=1).value == "dinesh.more@gmail.com" :
        for col in range(1, sheet.max_column+1):
            Dict[sheet.cell(row= 1, column=col).value] = sheet.cell(row= row, column=col).value


print(Dict)

